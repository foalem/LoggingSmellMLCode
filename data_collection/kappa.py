"""
Simplified kappa computation module

Behavior (simple and explicit):
1. Read two Excel files.
2. Find `snippet_id` column in each file (case-insensitive) and merge on it (inner join).
3. Find `Agreement` column in each file (case-insensitive) and use the values as-is.
4. Compute overall Cohen's kappa across all matched rows where both Agreement values are present.
5. If a smell/category column exists, compute Cohen's kappa per smell category (group by smell, treat NA as '<EMPTY>').
6. Write kappas to an Excel file (sheet: kappa_by_smell) and metadata to a second sheet. Write full-row disagreements to a separate Excel file and highlight them.

This implementation intentionally avoids normalization and other heuristics — it follows your instructions exactly.
"""
import os
import pandas as pd
from sklearn.metrics import cohen_kappa_score
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def _find_col_ci(df, candidates):
    """Find the first column in df whose lowercased name matches any candidate (case-insensitive).
    candidates may be a single name or an iterable of names.
    Returns the original column name or None.
    """
    if isinstance(candidates, str):
        candidates = [candidates]
    cols = {c.lower(): c for c in df.columns}
    for cand in candidates:
        cl = cand.lower()
        if cl in cols:
            return cols[cl]
    # substring match fallback
    for col_lower, col_orig in cols.items():
        for cand in candidates:
            if cand.lower() in col_lower:
                return col_orig
    return None


def compute_kappa(file_a, file_b, out_kappa, out_disagreements, snippet_id_col='snippet_id', agreement_col='Agreement', smell_candidates=('smell', 'smell_category', 'category'), min_samples=2):
    """Simplified kappa computation.

    Parameters
    - file_a, file_b: input Excel files
    - out_kappa: output Excel file path for kappa results
    - out_disagreements: output Excel file path for disagreements (highlighted)
    - snippet_id_col: name to look for the unique id column (case-insensitive)
    - agreement_col: name to look for Agreement column (case-insensitive)
    - smell_candidates: candidate names for smell/category column
    - min_samples: minimum samples to compute kappa for a group

    Returns (kappa_df, disagreements_df)
    """
    # read files
    if not os.path.exists(file_a):
        raise FileNotFoundError(file_a)
    if not os.path.exists(file_b):
        raise FileNotFoundError(file_b)

    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    # find snippet_id in both files
    sid_a = _find_col_ci(df_a, snippet_id_col)
    sid_b = _find_col_ci(df_b, snippet_id_col)
    if sid_a is None or sid_b is None:
        raise ValueError(f"Could not find '{snippet_id_col}' in both files (found: {sid_a}, {sid_b})")

    # find Agreement columns
    ag_a = _find_col_ci(df_a, agreement_col)
    ag_b = _find_col_ci(df_b, agreement_col)
    if ag_a is None or ag_b is None:
        raise ValueError(f"Could not find Agreement column in both files (found: {ag_a}, {ag_b})")

    # merge on snippet id (inner join)
    merged = pd.merge(df_a, df_b, left_on=sid_a, right_on=sid_b, suffixes=('_a', '_b'))
    if merged.empty:
        raise ValueError('No matching snippet_id rows after merge')

    # determine smell column if present
    smell_col = None
    s_a = _find_col_ci(df_a, smell_candidates)
    s_b = _find_col_ci(df_b, smell_candidates)
    # prefer the smell column that exists in the merged frame (with suffix if necessary)
    if s_a is not None and f"{s_a}_a" in merged.columns:
        smell_col = f"{s_a}_a"
    elif s_b is not None and f"{s_b}_b" in merged.columns:
        smell_col = f"{s_b}_b"
    elif s_a is not None and s_a in merged.columns:
        smell_col = s_a
    elif s_b is not None and s_b in merged.columns:
        smell_col = s_b

    # use agreement columns from merged (they may have suffixes)
    ag_col_merged_a = ag_a if ag_a in merged.columns else f"{ag_a}_a" if f"{ag_a}_a" in merged.columns else None
    ag_col_merged_b = ag_b if ag_b in merged.columns else f"{ag_b}_b" if f"{ag_b}_b" in merged.columns else None
    if ag_col_merged_a is None or ag_col_merged_b is None:
        # fallback: find any columns ending with _a/_b and containing "agree" in the name
        for c in merged.columns:
            if c.lower().endswith('_a') and 'agre' in c.lower():
                ag_col_merged_a = c
            if c.lower().endswith('_b') and 'agre' in c.lower():
                ag_col_merged_b = c
    if ag_col_merged_a is None or ag_col_merged_b is None:
        raise ValueError('Could not locate agreement columns in merged dataframe')

    # select only rows where both agreement values are present (non-null)
    mask_both = merged[ag_col_merged_a].notna() & merged[ag_col_merged_b].notna()
    data = merged[mask_both].copy()

    if data.empty:
        raise ValueError('No rows with both Agreement values present')

    results = []

    # overall kappa (ALL)
    a_vals = data[ag_col_merged_a]
    b_vals = data[ag_col_merged_b]
    n_all = len(data)
    overall_kappa = None
    if n_all >= min_samples:
        try:
            overall_kappa = cohen_kappa_score(a_vals, b_vals)
        except Exception:
            overall_kappa = None
    results.append({'smell': 'ALL', 'n_samples': int(n_all), 'kappa': overall_kappa})

    # per-smell kappas if smell column exists
    if smell_col is not None:
        # group by smell, treating NaN as '<EMPTY>'
        groups = data.groupby(data[smell_col].fillna('<EMPTY>'))
        for smell_val, grp in groups:
            a_g = grp[ag_col_merged_a]
            b_g = grp[ag_col_merged_b]
            n = len(grp)
            k = None
            if n >= min_samples:
                try:
                    k = cohen_kappa_score(a_g, b_g)
                except Exception:
                    k = None
            results.append({'smell': smell_val, 'n_samples': int(n), 'kappa': k})

    kappa_df = pd.DataFrame(results)

    # write kappa + metadata to Excel
    with pd.ExcelWriter(out_kappa, engine='openpyxl') as writer:
        kappa_df.to_excel(writer, sheet_name='kappa_by_smell', index=False)
        meta = {
            'file_a': file_a,
            'file_b': file_b,
            'rows_file_a': len(df_a),
            'rows_file_b': len(df_b),
            'rows_merged_after_join': len(merged),
            'rows_with_both_agreement': len(data)
        }
        pd.DataFrame([meta]).to_excel(writer, sheet_name='metadata', index=False)

    # disagreements: rows where the two Agreement columns differ
    dis_mask = data[ag_col_merged_a] != data[ag_col_merged_b]
    disagreements = data[dis_mask].copy()
    # save disagreements and highlight in Excel
    disagreements.to_excel(out_disagreements, index=False)
    if not disagreements.empty:
        wb = load_workbook(out_disagreements)
        ws = wb.active
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.fill = fill
        wb.save(out_disagreements)

    return kappa_df, disagreements


def cli_compute_kappa(args):
    file_a = args.file_a
    file_b = args.file_b
    out_kappa = args.out_kappa
    out_disagreements = args.out_disagreements
    min_samples = getattr(args, 'min_samples', 2)
    kdf, ddf = compute_kappa(file_a, file_b, out_kappa, out_disagreements, min_samples=min_samples)
    print(f'Wrote kappa results to: {out_kappa} (rows: {len(kdf)})')
    print(f'Wrote disagreement rows to: {out_disagreements} (rows: {len(ddf)})')
